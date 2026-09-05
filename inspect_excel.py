import openpyxl

def safe_val(cell):
    try:
        col = cell.column_letter
    except:
        col = '?'
    return (cell.value, col)

# Read computer book.xlsx
wb = openpyxl.load_workbook(r'computer book.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False)):
        vals = [safe_val(cell) for cell in row]
        print(f"Row {i+1}: {vals}")
wb.close()

print("\n\n" + "="*80)

# Read verification_4aug25.xlsx
wb2 = openpyxl.load_workbook(r'verification_4aug25.xlsx')
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False)):
        vals = [safe_val(cell) for cell in row]
        print(f"Row {i+1}: {vals}")
wb2.close()
